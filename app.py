# app.py
import eventlet # Necessário para o SocketIO rodar de forma assíncrona
eventlet.monkey_patch()

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file
from flask import g # 'g' é para armazenar a conexão temporariamente
from psycopg2.extras import RealDictCursor 
from flask_socketio import SocketIO, emit # Importacao necessária para WebSockets
import psycopg2
import hashlib
from datetime import date, timedelta
import io
import os

# --- NOVO IMPORT PARA EXCEL ---
import openpyxl 
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = "chave_secreta_super_segura_troque_em_producao"

# ----------------------------------------------------------------------
# INICIALIZAÇÃO DO SOCKETIO CORRIGIDA
# ----------------------------------------------------------------------
# Configurar o SocketIO
socketio = SocketIO(
    app,
    async_mode='eventlet', # Confirma que o worker assíncrono é o eventlet
    # Prioriza o 'polling' (mais estável no Gunicorn/Render) antes de tentar 'websocket'
    transports=['polling', 'websocket'] 
)
# ----------------------------------------------------------------------


def get_db():
    """Conecta ao banco de dados PostgreSQL e armazena em g."""
    if 'db' not in g:
        DATABASE_URL = os.environ.get('DATABASE_URL')
        if not DATABASE_URL:
            raise Exception("DATABASE_URL não configurada. Configure no Render ou localmente.")
        
        # Adiciona sslmode=require para compatibilidade com o Render
        g.db = psycopg2.connect(DATABASE_URL + "?sslmode=require")
    return g.db

def query_db(query, args=(), one=False, commit=False):
    """Executa uma query no PostgreSQL."""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor) 
    
    try:
        cur.execute(query, args)
        
        if commit:
            conn.commit()
            return None
        
        rows = cur.fetchall()
        return (rows[0] if rows else None) if one else rows
        
    except psycopg2.Error as e:
        conn.rollback()
        print(f"Erro no DB: {e}")
        raise
    finally:
        cur.close()

def generate_password_hash(password: str) -> str:
    return hashlib.sha256(password.encode('utf-8')).hexdigest()

def check_password_hash(stored_hash: str, password: str) -> bool:
    return stored_hash == hashlib.sha256(password.encode('utf-8')).hexdigest()

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"): return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

@app.teardown_appcontext
def close_db(e=None):
    """Fecha a conexão com o DB ao final do request."""
    db = g.pop('db', None)
    if db is not None:
        db.close()

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        
        # CORREÇÃO SQL: Trocando '?' por '%s'
        user = query_db("SELECT * FROM public.users WHERE username = %s", (username,), one=True)
        
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["is_admin"] = bool(user["is_admin"])
            return redirect(url_for("dashboard"))
        flash("Credenciais inválidas", "danger")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/")
@login_required
def dashboard():
    return render_template("dashboard.html", username=session.get("username"), is_admin=session.get("is_admin"))

# --- NOVA ROTA DE EXPORTAÇÃO EXCEL ---
@app.route("/export/excel")
@login_required
def export_excel():
    conn = get_db()
    
    # Busca dados das 4 tabelas 
    estoque = query_db("SELECT cod, descricao, unid, entradas, saidas, estoque_minimo, (entradas - saidas) as saldo FROM public.estoque ORDER BY cod")
    entradas = query_db("SELECT data, cod, descricao, unid, quantidade FROM public.entradas ORDER BY data DESC")
    saidas = query_db("SELECT data, cod, descricao, unid, quantidade FROM public.saidas ORDER BY data DESC")
    itens = query_db("SELECT cod, descricao, unid, estoque_minimo FROM public.itens ORDER BY cod")

    # Cria Workbook
    wb = openpyxl.Workbook()
    
    # Estilos Padrão (Omitido para brevidade, assumindo que funcionam)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="192a56", end_color="192a56", fill_type="solid")
    center_align = Alignment(horizontal='center')

    def style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

    # --- ABA 1: Visão Geral (Estoque) ---
    ws1 = wb.active
    ws1.title = "Visão Geral"
    ws1.append(["Código", "Descrição", "Unidade", "Mínimo", "Entradas", "Saídas", "Saldo Atual", "Status"])
    style_header(ws1)
    
    for row in estoque:
        # Acesso por nome de coluna graças ao RealDictCursor no query_db
        status = "BAIXO" if row['saldo'] <= row['estoque_minimo'] else "OK"
        ws1.append([row['cod'], row['descricao'], row['unid'], row['estoque_minimo'], row['entradas'], row['saidas'], row['saldo'], status])
        
        # Pinta de vermelho se estiver baixo
        if status == "BAIXO":
            ws1.cell(row=ws1.max_row, column=8).font = Font(color="FF0000", bold=True)
        else:
            ws1.cell(row=ws1.max_row, column=8).font = Font(color="008000", bold=True)

    # --- ABA 2: Histórico Entradas ---
    ws2 = wb.create_sheet("Entradas")
    ws2.append(["Data", "Código", "Descrição", "Unidade", "Quantidade"])
    style_header(ws2)
    for row in entradas:
        ws2.append([row['data'], row['cod'], row['descricao'], row['unid'], row['quantidade']])

    # --- ABA 3: Histórico Saídas ---
    ws3 = wb.create_sheet("Saídas")
    ws3.append(["Data", "Código", "Descrição", "Unidade", "Quantidade"])
    style_header(ws3)
    for row in saidas:
        ws3.append([row['data'], row['cod'], row['descricao'], row['unid'], row['quantidade']])

    # --- ABA 4: Cadastros Base ---
    ws4 = wb.create_sheet("Cadastros")
    ws4.append(["Código", "Descrição", "Unidade", "Estoque Mínimo Padrão"])
    style_header(ws4)
    for row in itens:
        ws4.append([row['cod'], row['descricao'], row['unid'], row['estoque_minimo']])

    # Ajuste largura das colunas
    for ws in wb.worksheets:
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['A'].width = 15

    # Salva em memória
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"Relatorio_Estoque_{date.today().strftime('%d-%m-%Y')}.xlsx"
    return send_file(out, download_name=filename, as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --- APIS DO DASHBOARD ---
@app.route("/api/dashboard/stats")
@login_required
def api_stats():
    total_itens = query_db("SELECT COUNT(*) as c FROM public.itens", one=True)['c']
    baixa = query_db("SELECT COUNT(*) as c FROM public.estoque WHERE (entradas - saidas) <= estoque_minimo", one=True)['c']
    hoje = date.today().isoformat()
    
    # CORREÇÃO SQL: Trocando '?' por '%s'
    mov_ent = query_db("SELECT COUNT(*) as c FROM public.entradas WHERE data = %s", (hoje,), one=True)['c']
    mov_sai = query_db("SELECT COUNT(*) as c FROM public.saidas WHERE data = %s", (hoje,), one=True)['c']
    
    chart_labels = []
    chart_ent = []
    chart_sai = []
    for i in range(6, -1, -1):
        d = (date.today() - timedelta(days=i)).isoformat()
        chart_labels.append(d.split('-')[2] + '/' + d.split('-')[1])
        # CORREÇÃO SQL: Trocando '?' por '%s'
        qe = query_db("SELECT SUM(quantidade) as q FROM public.entradas WHERE data = %s", (d,), one=True)['q'] or 0
        qs = query_db("SELECT SUM(quantidade) as q FROM public.saidas WHERE data = %s", (d,), one=True)['q'] or 0
        chart_ent.append(qe)
        chart_sai.append(qs)

    return jsonify({
        "total_itens": total_itens, "alertas": baixa, "mov_hoje": mov_ent + mov_sai,
        "chart": { "labels": chart_labels, "entrada": chart_ent, "saida": chart_sai }
    })

@app.route("/api/estoque")
@login_required
def api_estoque():
    rows = query_db("SELECT cod, descricao, unid, entradas, saidas, estoque_minimo FROM public.estoque ORDER BY cod")
    results = []
    for r in rows:
        d = dict(r)
        d['saldo'] = d['entradas'] - d['saidas']
        d['alerta_baixo'] = d['saldo'] <= (d['estoque_minimo'] or 0)
        results.append(d)
    return jsonify(results)

@app.route("/api/itens", methods=["GET", "POST"])
@login_required
def api_itens_handler():
    if request.method == 'GET':
        return jsonify([dict(r) for r in query_db("SELECT * FROM public.itens ORDER BY cod")])
    data = request.json
    cod = data.get("cod", "").strip().upper()
    
    # CORREÇÃO SQL: Trocando '?' por '%s'
    if query_db("SELECT 1 FROM public.itens WHERE cod=%s", (cod,), one=True): return jsonify({"error": "Código já existe"}), 400
    
    # CORREÇÃO SQL: Trocando '?' por '%s' e adicionando public.
    query_db("INSERT INTO public.itens (cod, descricao, unid, estoque_minimo) VALUES (%s, %s, %s, %s)", 
             (cod, data['descricao'], data['unid'], int(data.get("estoque_minimo", 10))), commit=True)
    query_db("INSERT INTO public.estoque (cod, descricao, unid, estoque_minimo) VALUES (%s, %s, %s, %s)", 
             (cod, data['descricao'], data['unid'], int(data.get("estoque_minimo", 10))), commit=True)
    
    # EMITIR EVENTO APÓS CADASTRO DE NOVO ITEM
    socketio.emit('estoque_atualizado', {'message': 'Novo item cadastrado.'})
    
    return jsonify({"ok": True})

@app.route("/api/entrada", methods=["POST"])
@login_required
def api_entrada():
    d = request.json
    
    # CORREÇÃO SQL: Trocando '?' por '%s'
    item = query_db("SELECT * FROM public.estoque WHERE cod=%s", (d['cod'],), one=True)
    
    if not item: return jsonify({"error": "Item não encontrado"}), 404
    
    # CORREÇÃO SQL: Trocando '?' por '%s' e adicionando public.
    query_db("INSERT INTO public.entradas (cod, descricao, unid, quantidade, data) VALUES (%s, %s, %s, %s, %s)", 
             (d['cod'], item['descricao'], item['unid'], int(d['qtd']), d['data']), commit=True)
             
    # CORREÇÃO SQL: Trocando '?' por '%s'
    query_db("UPDATE public.estoque SET entradas = entradas + %s WHERE cod=%s", (int(d['qtd']), d['cod']), commit=True)
    
    # EMITIR EVENTO DE TEMPO REAL
    socketio.emit('estoque_atualizado', {'message': 'Entrada registrada. Recarregando dados.'})
    
    return jsonify({"ok": True})

@app.route("/api/saida", methods=["POST"])
@login_required
def api_saida():
    d = request.json
    
    # CORREÇÃO SQL: Trocando '?' por '%s'
    item = query_db("SELECT * FROM public.estoque WHERE cod=%s", (d['cod'],), one=True)
    
    if not item: return jsonify({"error": "Item não encontrado"}), 404
    saldo = item['entradas'] - item['saidas']
    if int(d['qtd']) > saldo: return jsonify({"error": f"Saldo insuficiente ({saldo})"}), 400
    
    # CORREÇÃO SQL: Trocando '?' por '%s' e adicionando public.
    query_db("INSERT INTO public.saidas (cod, descricao, unid, quantidade, data) VALUES (%s, %s, %s, %s, %s)", 
             (d['cod'], item['descricao'], item['unid'], int(d['qtd']), d['data']), commit=True)
             
    # CORREÇÃO SQL: Trocando '?' por '%s'
    query_db("UPDATE public.estoque SET saidas = saidas + %s WHERE cod=%s", (int(d['qtd']), d['cod']), commit=True)
    
    # EMITIR EVENTO DE TEMPO REAL
    socketio.emit('estoque_atualizado', {'message': 'Saída registrada. Recarregando dados.'})
    
    return jsonify({"ok": True})

@app.route("/api/movimentacoes")
@login_required
def api_movimentacoes():
    ent = query_db("SELECT * FROM public.entradas ORDER BY data DESC, id DESC LIMIT 20")
    sai = query_db("SELECT * FROM public.saidas ORDER BY data DESC, id DESC LIMIT 20")
    return jsonify({"entradas": [dict(r) for r in ent], "saidas": [dict(r) for r in sai]})

# Users (Admin)
@app.route("/api/users", methods=["GET", "POST"])
@login_required
def api_users():
    if not session.get("is_admin"): return jsonify({"error": "Acesso negado"}), 403
    if request.method == 'GET':
        return jsonify([dict(r) for r in query_db("SELECT id, username, is_admin FROM public.users")])
    d = request.json
    try:
        # CORREÇÃO SQL: Trocando '?' por '%s' e adicionando public.
        query_db("INSERT INTO public.users (username, password_hash, is_admin) VALUES (%s, %s, %s)", 
                 (d['username'], generate_password_hash(d['password']), 1 if d['is_admin'] else 0), commit=True)
        return jsonify({"ok": True})
    except: return jsonify({"error": "Erro/Duplicado"}), 400

@app.route("/api/users/<int:uid>", methods=["DELETE"])
@login_required
def api_del_user(uid):
    if not session.get("is_admin"): return jsonify({"error": "Acesso negado"}), 403
    if uid == session["user_id"]: return jsonify({"error": "Não delete a si mesmo"}), 400
    
    # CORREÇÃO SQL: Trocando '?' por '%s'
    query_db("DELETE FROM public.users WHERE id=%s", (uid,), commit=True)
    return jsonify({"ok": True})

# --- Execução Principal (Adaptado para SocketIO/eventlet) ---
if __name__ == "__main__":
    # Usa a porta definida pelo ambiente (Render/Heroku) ou 5002 localmente
    port = int(os.environ.get("PORT", 5002))
    
    # Rodar o SocketIO com eventlet
    # O Gunicorn (no Render) deve ser configurado separadamente para usar --worker-class eventlet
    print(f"Iniciando SocketIO na porta {port}...")
    socketio.run(app, debug=True, host="0.0.0.0", port=port, allow_unsafe_werkzeug=True)


